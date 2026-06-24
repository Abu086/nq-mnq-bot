"""
NIFTY / BANKNIFTY Paper Trading Bot — 4 Strategy Comparison
=============================================================
Strategies run SIMULTANEOUSLY every day for comparison:
  1. Bull Call Spread  — always trades bullish side
  2. Bear Put Spread   — always trades bearish side
  3. Both              — runs both every day regardless of trend
  4. Smart Trend       — picks ONE based on EMA9 vs EMA21 + price signal

Entry Time  : 9:20 AM IST
Expiry      : NIFTY → Weekly (Thursday) | BANKNIFTY → Monthly (last Thursday)
Budget      : ₹50,000 per strategy per instrument
Journal     : trading_journal.xlsx (8 strategy sheets + summary)
Mode        : PAPER TRADING — NO REAL ORDERS
"""

import datetime
import json
import logging
import math
import os
import random
import sys
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from angel_one_client import AngelOneClient, get_live_index_price, get_option_ltp
    ANGEL_AVAILABLE = True
except ImportError:
    ANGEL_AVAILABLE = False

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s IST] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("logs/india_bot.log"),
    ],
)
log = logging.getLogger(__name__)

IST          = ZoneInfo("Asia/Kolkata")
JOURNAL_FILE = "trading_journal.xlsx"

# ── Instrument Config ─────────────────────────────────────────────────────────
INSTRUMENTS = {
    "NIFTY": {
        "ticker":       "^NSEI",
        "lot_size":     65,   # Revised from 75 to 65, effective Jan 2026 per NSE circular FAOP70616
        "strike_width": 100,
        "budget":       50000,
        "expiry_type":  "weekly",
        "stt":          0.0125,
        "sebi_fee":     0.0001,
        "stamp_duty":   0.003,
        "brokerage":    20,
        "gst_pct":      0.18,
    },
    "BANKNIFTY": {
        "ticker":       "^NSEBANK",
        "lot_size":     30,
        "strike_width": 200,
        "budget":       50000,
        "expiry_type":  "monthly",
        "stt":          0.0125,
        "sebi_fee":     0.0001,
        "stamp_duty":   0.003,
        "brokerage":    20,
        "gst_pct":      0.18,
    },
}

# 4 strategies — each runs independently
STRATEGIES = [
    {"key": "BullCall", "label": "Bull Call Spread",       "mode": "bull_only"},
    {"key": "BearPut",  "label": "Bear Put Spread",        "mode": "bear_only"},
    {"key": "Both",     "label": "Both Spreads",           "mode": "both"},
    {"key": "Smart",    "label": "Smart Trend",            "mode": "smart"},
]

TARGET_PCT = 0.30
SL_PCT     = 0.30

# ── Holidays ──────────────────────────────────────────────────────────────────
INDIA_HOLIDAYS_2026 = {
    datetime.date(2026,  1, 26), datetime.date(2026,  2, 26),
    datetime.date(2026,  3, 17), datetime.date(2026,  4,  2),
    datetime.date(2026,  4, 14), datetime.date(2026,  4, 30),
    datetime.date(2026,  8, 15), datetime.date(2026,  8, 27),
    datetime.date(2026, 10,  2), datetime.date(2026, 10, 20),
    datetime.date(2026, 11,  9), datetime.date(2026, 11, 10),
    datetime.date(2026, 11, 25), datetime.date(2026, 12, 25),
}

def is_india_market_day(dt):
    return dt.weekday() < 5 and dt not in INDIA_HOLIDAYS_2026

# ── Expiry ────────────────────────────────────────────────────────────────────
def get_weekly_expiry(today):
    if today.weekday() == 3 and today not in INDIA_HOLIDAYS_2026:
        return today
    days = (3 - today.weekday()) % 7 or 7
    return today + datetime.timedelta(days=days)

def get_monthly_expiry(today):
    def last_thu(y, m):
        d = datetime.date(y + (m == 12), 1 if m == 12 else m + 1, 1) - datetime.timedelta(days=1)
        while d.weekday() != 3:
            d -= datetime.timedelta(days=1)
        return d
    exp = last_thu(today.year, today.month)
    if today > exp:
        exp = last_thu(today.year + (today.month == 12), 1 if today.month == 12 else today.month + 1)
    return exp

def get_expiry(inst_name, today):
    return get_weekly_expiry(today) if INSTRUMENTS[inst_name]["expiry_type"] == "weekly" else get_monthly_expiry(today)

# ── EMA ───────────────────────────────────────────────────────────────────────
def ema(prices, period):
    if len(prices) < period:
        return sum(prices) / len(prices)
    k   = 2 / (period + 1)
    val = sum(prices[:period]) / period
    for p in prices[period:]:
        val = p * k + val * (1 - k)
    return round(val, 2)

# ── Price Fetch ───────────────────────────────────────────────────────────────
def fetch_price_history(ticker):
    import urllib.request
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=30d"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())
        closes = [c for c in data["chart"]["result"][0]["indicators"]["quote"][0]["close"] if c]
        return {"current": round(closes[-1], 2), "prev_close": round(closes[-2], 2), "closes": closes}
    except:
        return None

# ── Trend Detection ───────────────────────────────────────────────────────────
def detect_trend(price_data):
    closes = price_data["closes"]
    cur, prev = price_data["current"], price_data["prev_close"]
    e9, e21   = ema(closes, 9), ema(closes, 21)
    ps = "BULLISH" if cur > prev else "BEARISH"
    es = "BULLISH" if e9 > e21  else "BEARISH"
    if ps == es == "BULLISH":
        return {"trend": "BULLISH", "signal": "BULL_CALL",
                "reason": f"Price↑({cur}>{prev}) & EMA9↑({e9}>{e21})", "ema9": e9, "ema21": e21}
    elif ps == es == "BEARISH":
        return {"trend": "BEARISH", "signal": "BEAR_PUT",
                "reason": f"Price↓({cur}<{prev}) & EMA9↓({e9}<{e21})", "ema9": e9, "ema21": e21}
    else:
        return {"trend": "SKIP", "signal": None,
                "reason": f"Conflicting — Price:{ps} EMA:{es}", "ema9": e9, "ema21": e21}


# ── Angel One Live Option Premium Fetcher ─────────────────────────────────────
_angel_client = None

def get_angel_client():
    """Lazily create and login the Angel One client once per run."""
    global _angel_client
    if not ANGEL_AVAILABLE:
        return None
    if _angel_client is not None:
        return _angel_client
    try:
        client = AngelOneClient()
        if client.login():
            _angel_client = client
            return client
    except Exception as e:
        log.warning(f"Angel One client init/login failed: {e}")
    return None


def build_option_symbol(inst_name: str, expiry: datetime.date, strike: int, opt_type: str) -> str:
    """
    Build Angel One style trading symbol, e.g. NIFTY16JUN2623300CE
    Format: SYMBOL + DD + MON + YY + STRIKE + CE/PE
    """
    day = expiry.strftime("%d").upper()
    mon = expiry.strftime("%b").upper()
    yr  = expiry.strftime("%y")
    return f"{inst_name}{day}{mon}{yr}{strike}{opt_type}"


def fetch_real_option_premium(client, inst_name: str, expiry: datetime.date,
                              strike: int, opt_type: str):
    """
    Search for the option contract and fetch its live LTP.
    Returns (premium, trading_symbol, symbol_token) — premium is None on failure.
    """
    trading_symbol = build_option_symbol(inst_name, expiry, strike, opt_type)
    if client is None:
        return None, trading_symbol, None
    try:
        results = client.search_scrip("NFO", trading_symbol)
        if not results:
            log.warning(f"  No scrip found for {trading_symbol}")
            return None, trading_symbol, None
        token = results[0]["symboltoken"]
        ltp   = get_option_ltp(client, trading_symbol, token)
        return ltp, trading_symbol, token
    except Exception as e:
        log.warning(f"  Option premium fetch failed for {trading_symbol}: {e}")
        return None, trading_symbol, None


# ── Fees ──────────────────────────────────────────────────────────────────────
def calc_fees(inst, lots, premium):
    units  = lots * inst["lot_size"]
    to     = premium * units
    brk    = inst["brokerage"] * 2
    sebi   = inst["sebi_fee"] / 100 * to * 2
    stt    = inst["stt"] / 100 * to
    stamp  = inst["stamp_duty"] / 100 * to
    gst    = inst["gst_pct"] * (brk + sebi)
    return {
        "brokerage": round(brk, 2), "stt": round(stt, 2),
        "sebi_fee":  round(sebi,2), "stamp_duty": round(stamp, 2),
        "gst":       round(gst, 2), "total_fees": round(brk+sebi+stt+stamp+gst, 2),
    }

def get_atm(price, width):
    return int(math.floor(price / width) * width)

# ── Simulate One Spread ───────────────────────────────────────────────────────
def simulate_spread(inst_name, inst, price_data, spread_type, trend_info, strategy_label):
    """
    spread_type: 'BULL_CALL' or 'BEAR_PUT'
    Returns a trade dict or None if skipped.
    """
    today  = datetime.date.today()
    expiry = get_expiry(inst_name, today)
    dte    = (expiry - today).days
    width  = inst["strike_width"]
    budget = inst["budget"]
    spot   = price_data["current"]
    prev   = price_data["prev_close"]
    atm    = get_atm(spot, width)

    if spread_type == "BULL_CALL":
        buy_s, sell_s, opt = atm, atm + width, "CE"
        slabel = "Bull Call Spread"
    else:
        buy_s, sell_s, opt = atm + width, atm, "PE"
        slabel = "Bear Put Spread"

    # ── Try to fetch REAL live premiums from Angel One ──────────────────────
    client = get_angel_client()
    buy_ltp, buy_symbol, _   = fetch_real_option_premium(client, inst_name, expiry, buy_s, opt)
    sell_ltp, sell_symbol, _ = fetch_real_option_premium(client, inst_name, expiry, sell_s, opt)

    data_source = "SIMULATED"
    if buy_ltp is not None and sell_ltp is not None and buy_ltp > sell_ltp:
        # Real net debit premium for the spread = buy leg LTP - sell leg LTP
        entry_prem  = round(buy_ltp - sell_ltp, 2)
        data_source = "LIVE (Angel One)"
        log.info(f"  ✅ LIVE premiums — Buy {buy_symbol}: ₹{buy_ltp}  Sell {sell_symbol}: ₹{sell_ltp}")
    else:
        # Fallback — simulate premium based on DTE (Angel One data unavailable)
        if dte <= 2:   ep = random.uniform(0.15, 0.25)
        elif dte <= 7: ep = random.uniform(0.25, 0.40)
        elif dte <=15: ep = random.uniform(0.35, 0.50)
        else:          ep = random.uniform(0.40, 0.55)
        entry_prem = round(width * ep, 2)
        log.warning(f"  ⚠️  Using SIMULATED premium (Angel One data unavailable)")
    fees_est   = inst["brokerage"] * 2 + inst["stt"] / 100 * entry_prem * inst["lot_size"]
    lots       = max(1, int(budget / (entry_prem * inst["lot_size"] + fees_est)))
    units      = lots * inst["lot_size"]
    total_cost = round(entry_prem * units, 2)

    max_p_pu        = width - entry_prem
    max_p_tot       = round(max_p_pu * units, 2)
    total_cost_paid = round(entry_prem * units, 2)
    tgt             = round(total_cost_paid * TARGET_PCT, 2)    # 30% of total cost paid
    sl_amt          = round(total_cost_paid * SL_PCT, 2)        # 30% of total cost paid
    tgt_exit        = round(entry_prem * (1 + TARGET_PCT), 2)
    sl_exit         = round(entry_prem * (1 - SL_PCT), 2)

    r = random.random()
    win_rate = 0.60 if trend_info["trend"] != "SKIP" else 0.50
    if r < win_rate:
        outcome, exit_p, exit_m = "PROFIT",    tgt_exit, random.randint(20, 180)
    elif r < win_rate + 0.25:
        outcome, exit_p, exit_m = "STOP LOSS", sl_exit,  random.randint(10, 120)
    else:
        outcome, exit_p, exit_m = "FLAT", round(entry_prem * random.uniform(0.95,1.05),2), random.randint(240,360)

    gross   = round((exit_p - entry_prem) * units, 2)
    fees    = calc_fees(inst, lots, entry_prem)
    net     = round(gross - fees["total_fees"], 2)
    cap_use = round(total_cost + fees["total_fees"], 2)
    ret_pct = round(gross / total_cost_paid * 100, 1) if total_cost_paid else 0

    now_ist = datetime.datetime.now(IST)
    entry_t = now_ist.replace(hour=9, minute=20, second=0, microsecond=0)
    exit_t  = entry_t + datetime.timedelta(minutes=exit_m)

    return {
        "date":          now_ist.strftime("%Y-%m-%d"),
        "instrument":    inst_name,
        "strategy_mode": strategy_label,
        "spread":        slabel,
        "data_source":   data_source,
        "trend":         trend_info["trend"],
        "trend_reason":  trend_info["reason"],
        "ema9":          trend_info["ema9"],
        "ema21":         trend_info["ema21"],
        "expiry":        expiry.strftime("%Y-%m-%d"),
        "expiry_type":   inst["expiry_type"].upper(),
        "dte":           dte,
        "entry_time":    entry_t.strftime("%H:%M IST"),
        "exit_time":     exit_t.strftime("%H:%M IST"),
        "spot_at_entry": spot,
        "prev_close":    prev,
        "buy_strike":    f"{buy_s} {opt}",
        "sell_strike":   f"{sell_s} {opt}",
        "lots":          lots,
        "lot_size":      inst["lot_size"],
        "units":         units,
        "entry_premium": entry_prem,
        "exit_premium":  exit_p,
        "total_cost":    total_cost,
        "capital_used":  cap_use,
        "max_profit":    max_p_tot,
        "target":        tgt,
        "stop_loss":     sl_amt,
        "gross_pnl":     gross,
        "return_pct":    ret_pct,
        "brokerage":     fees["brokerage"],
        "stt":           fees["stt"],
        "sebi_fee":      fees["sebi_fee"],
        "stamp_duty":    fees["stamp_duty"],
        "gst":           fees["gst"],
        "total_fees":    fees["total_fees"],
        "net_pnl":       net,
        "outcome":       outcome,
        "budget":        budget,
        "mode":          "PAPER TRADING",
    }

# ── Excel ─────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
PRF_FILL  = PatternFill("solid", start_color="C6EFCE")
SL_FILL   = PatternFill("solid", start_color="FFC7CE")
FLT_FILL  = PatternFill("solid", start_color="FFEB9C")
SKP_FILL  = PatternFill("solid", start_color="E2EFDA")
WHT_FILL  = PatternFill("solid", start_color="FFFFFF")
SUB_FILL  = PatternFill("solid", start_color="D6E4F0")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))

COLS = [
    ("Date",             "date",           13),
    ("Instrument",       "instrument",     13),
    ("Strategy Mode",    "strategy_mode",  18),
    ("Spread",           "spread",         20),
    ("Data Source",      "data_source",    16),
    ("Trend",            "trend",          11),
    ("Trend Reason",     "trend_reason",   35),
    ("EMA 9",            "ema9",           10),
    ("EMA 21",           "ema21",          10),
    ("Expiry",           "expiry",         13),
    ("Expiry Type",      "expiry_type",    12),
    ("DTE",              "dte",             7),
    ("Entry Time",       "entry_time",     12),
    ("Exit Time",        "exit_time",      12),
    ("Spot at Entry",    "spot_at_entry",  15),
    ("Prev Close",       "prev_close",     12),
    ("Buy Strike",       "buy_strike",     14),
    ("Sell Strike",      "sell_strike",    14),
    ("Lots",             "lots",            8),
    ("Lot Size",         "lot_size",        9),
    ("Units",            "units",           9),
    ("Entry Premium",    "entry_premium",  15),
    ("Exit Premium",     "exit_premium",   14),
    ("Total Cost (₹)",   "total_cost",     14),
    ("Capital Used (₹)", "capital_used",   15),
    ("Max Profit (₹)",   "max_profit",     14),
    ("Target (₹)",       "target",         13),
    ("Stop Loss (₹)",    "stop_loss",      13),
    ("Gross P&L (₹)",    "gross_pnl",      13),
    ("Return %",         "return_pct",     12),
    ("Brokerage (₹)",    "brokerage",      13),
    ("STT (₹)",          "stt",            10),
    ("SEBI Fee (₹)",     "sebi_fee",       12),
    ("Stamp Duty (₹)",   "stamp_duty",     13),
    ("GST (₹)",          "gst",            10),
    ("Total Fees (₹)",   "total_fees",     13),
    ("Net P&L (₹)",      "net_pnl",        13),
    ("Outcome",          "outcome",        12),
    ("Budget (₹)",       "budget",         12),
    ("Mode",             "mode",           22),
]

# Sheet title colours per strategy
STRAT_COLORS = {
    "BullCall": "1A5276",
    "BearPut":  "922B21",
    "Both":     "1E8449",
    "Smart":    "6C3483",
}

def _hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill      = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = THIN

def _data(ws, row, col, val, outcome=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = THIN
    c.fill = {"PROFIT": PRF_FILL, "STOP LOSS": SL_FILL,
              "FLAT": FLT_FILL, "SKIP": SKP_FILL}.get(outcome, WHT_FILL)

def build_sheet(wb, sheet_name, title, color_key):
    if sheet_name in wb.sheetnames:
        return
    ws    = wb.create_sheet(sheet_name)
    color = STRAT_COLORS.get(color_key, "0D2137")
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    t       = ws["A1"]
    t.value = title
    t.font  = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    t.fill  = PatternFill("solid", start_color=color)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for i, (h, _, w) in enumerate(COLS, 1):
        _hdr(ws, 2, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 35
    ws.freeze_panes = "A3"

def ensure_all_sheets(wb):
    for inst in ["NIFTY", "BANKNIFTY"]:
        for s in STRATEGIES:
            sname = f"{inst}_{s['key']}"
            title = f"{inst} — {s['label']} | Paper Trading"
            build_sheet(wb, sname, title, s["key"])
    # Skipped days sheet
    if "Skipped_Days" not in wb.sheetnames:
        ws = wb.create_sheet("Skipped_Days")
        ws.merge_cells("A1:G1")
        t       = ws["A1"]
        t.value = "Skipped Days — Smart Strategy Conflicting Signals"
        t.font  = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        t.fill  = PatternFill("solid", start_color="0D2137")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        for i, (h, w) in enumerate([
            ("Date",12),("Instrument",13),("Trend",11),
            ("Reason",45),("EMA9",10),("EMA21",10),("Action",25)
        ], 1):
            _hdr(ws, 2, i, h)
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"

def append_trade(wb, trade, sheet_name):
    ws  = wb[sheet_name]
    row = ws.max_row + 1
    out = trade.get("outcome", "")
    for i, (_, key, _) in enumerate(COLS, 1):
        _data(ws, row, i, trade.get(key, ""), out)

def append_skip(wb, inst_name, trend_info):
    ws  = wb["Skipped_Days"]
    row = ws.max_row + 1
    now = datetime.datetime.now(IST)
    for col, val in enumerate([
        now.strftime("%Y-%m-%d"), inst_name,
        trend_info["trend"], trend_info["reason"],
        trend_info["ema9"], trend_info["ema21"],
        "NO TRADE — Conflicting signals"
    ], 1):
        _data(ws, row, col, val, "SKIP")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 65)
    log.info("  NIFTY/BANKNIFTY — 4 STRATEGY COMPARISON BOT")
    log.info("=" * 65)

    today = datetime.date.today()
    if not is_india_market_day(today):
        log.info(f"{today} is not a trading day. Exiting.")
        return

    if not os.path.exists(JOURNAL_FILE):
        log.error("Journal not found. Run trading_bot.py first.")
        return

    wb = openpyxl.load_workbook(JOURNAL_FILE)
    ensure_all_sheets(wb)

    all_trades = []

    for inst_name, inst in INSTRUMENTS.items():
        log.info(f"\n{'─'*55}")
        log.info(f"  {inst_name}")
        log.info(f"{'─'*55}")

        price_data = fetch_price_history(inst["ticker"])
        if price_data is None:
            base = 23350.0 if inst_name == "NIFTY" else 50200.0
            closes = [base + random.uniform(-200, 200) for _ in range(25)]
            price_data = {"current": round(closes[-1],2),
                          "prev_close": round(closes[-2],2), "closes": closes}
            log.warning("  Using simulated prices")

        log.info(f"  Current    : {price_data['current']:,.2f}")
        log.info(f"  Prev Close : {price_data['prev_close']:,.2f}")

        trend_info = detect_trend(price_data)
        log.info(f"  Trend      : {trend_info['trend']}  ({trend_info['reason']})")

        expiry = get_expiry(inst_name, today)
        log.info(f"  Expiry     : {expiry}  DTE: {(expiry-today).days}")

        for strat in STRATEGIES:
            sheet_name = f"{inst_name}_{strat['key']}"
            mode       = strat["mode"]
            log.info(f"\n  [{strat['label']}]")

            if mode == "bull_only":
                t = simulate_spread(inst_name, inst, price_data, "BULL_CALL", trend_info, strat["label"])
                append_trade(wb, t, sheet_name)
                all_trades.append((sheet_name, t))
                log.info(f"  Outcome: {t['outcome']}  Net: ₹{t['net_pnl']:,.2f}")

            elif mode == "bear_only":
                t = simulate_spread(inst_name, inst, price_data, "BEAR_PUT", trend_info, strat["label"])
                append_trade(wb, t, sheet_name)
                all_trades.append((sheet_name, t))
                log.info(f"  Outcome: {t['outcome']}  Net: ₹{t['net_pnl']:,.2f}")

            elif mode == "both":
                for stype in ["BULL_CALL", "BEAR_PUT"]:
                    t = simulate_spread(inst_name, inst, price_data, stype, trend_info, strat["label"])
                    append_trade(wb, t, sheet_name)
                    all_trades.append((sheet_name, t))
                    log.info(f"  {t['spread']:22} Outcome: {t['outcome']}  Net: ₹{t['net_pnl']:,.2f}")

            elif mode == "smart":
                if trend_info["signal"] is None:
                    log.info(f"  ⚠️  SKIP — conflicting signals")
                    append_skip(wb, inst_name, trend_info)
                else:
                    t = simulate_spread(inst_name, inst, price_data,
                                        trend_info["signal"], trend_info, strat["label"])
                    append_trade(wb, t, sheet_name)
                    all_trades.append((sheet_name, t))
                    log.info(f"  {t['spread']:22} Outcome: {t['outcome']}  Net: ₹{t['net_pnl']:,.2f}")

    wb.save(JOURNAL_FILE)
    log.info(f"\n✅ Journal saved → {JOURNAL_FILE}")

    # Summary
    log.info("\n" + "=" * 65)
    log.info("  SUMMARY — ALL STRATEGIES")
    log.info("=" * 65)
    log.info(f"  {'Sheet':<25} {'Spread':<22} {'Outcome':<12} {'Net P&L':>12}")
    log.info(f"  {'─'*72}")
    total = 0
    for sname, t in all_trades:
        log.info(f"  {sname:<25} {t['spread']:<22} {t['outcome']:<12} ₹{t['net_pnl']:>10,.2f}")
        total += t["net_pnl"]
    log.info(f"  {'─'*72}")
    log.info(f"  {'TOTAL':<60} ₹{total:>10,.2f}")
    log.info("=" * 65)

if __name__ == "__main__":
    main()
