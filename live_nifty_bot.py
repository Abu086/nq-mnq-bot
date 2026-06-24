"""
NIFTY LIVE Trading Bot — Bull Call Spread
==========================================
- Trades REAL money via Angel One SmartAPI
- Entry: 9:20 AM IST every day EXCEPT expiry day
- Expiry day = Tuesday (or previous working day if Tuesday is holiday)
- Strategy: Bull Call Spread ONLY when trend is Bullish (EMA9 > EMA21 + Price > Prev Close)
- Budget: Rs.15,000 per trade
- Target: 30% of total cost paid
- Stop Loss: 30% of total cost paid
- Contract: Always uses nearest upcoming Tuesday expiry
- Monitors position every 5 minutes until target/SL hit or 3:15 PM IST
- Logs ALL trades to trading_journal.xlsx (appends, never overwrites)
- Mode: LIVE TRADING — REAL MONEY
"""

import datetime
import json
import logging
import math
import os
import sys
import time
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from angel_one_client import AngelOneClient, generate_totp

# ── Logging ───────────────────────────────────────────────────────────────────
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s IST] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("logs/live_nifty.log"),
    ],
)
log = logging.getLogger(__name__)

IST          = ZoneInfo("Asia/Kolkata")
JOURNAL_FILE = "trading_journal.xlsx"

# ── Config ────────────────────────────────────────────────────────────────────
BUDGET       = 15000       # Rs.15,000 live budget
TARGET_PCT   = 0.30        # 30% of cost paid
SL_PCT       = 0.30        # 30% of cost paid
LOT_SIZE     = 65          # NIFTY lot size (revised from 75 to 65, effective Jan 2026 per NSE circular FAOP70616)
STRIKE_WIDTH = 100         # 100 point spread
POLL_MINS    = 5           # Check position every 5 minutes
ENTRY_HOUR   = 9
ENTRY_MIN    = 20
EXIT_HOUR    = 15
EXIT_MIN     = 15

# ── Holiday List 2026 ─────────────────────────────────────────────────────────
INDIA_HOLIDAYS_2026 = {
    datetime.date(2026,  1, 26), datetime.date(2026,  2, 26),
    datetime.date(2026,  3, 17), datetime.date(2026,  4,  2),
    datetime.date(2026,  4, 14), datetime.date(2026,  4, 30),
    datetime.date(2026,  8, 15), datetime.date(2026,  8, 27),
    datetime.date(2026, 10,  2), datetime.date(2026, 10, 20),
    datetime.date(2026, 11,  9), datetime.date(2026, 11, 10),
    datetime.date(2026, 11, 25), datetime.date(2026, 12, 25),
}

def get_weekly_expiry(today: datetime.date) -> datetime.date:
    """
    Get this week's NIFTY expiry date.
    Normally = Tuesday of current week.
    If Tuesday is a holiday → shift to previous working day.
    """
    # Find this week's Tuesday
    days_to_tuesday = (1 - today.weekday()) % 7
    tuesday = today + datetime.timedelta(days=days_to_tuesday)

    # If today IS Tuesday, use today's Tuesday
    if today.weekday() == 1:
        tuesday = today

    # If Tuesday is a holiday, shift to previous working day
    while tuesday in INDIA_HOLIDAYS_2026 or tuesday.weekday() >= 5:
        tuesday -= datetime.timedelta(days=1)

    return tuesday


def get_trade_expiry(today: datetime.date) -> datetime.date:
    """
    Get the expiry contract to trade today.
    - If today is Monday → use THIS week's expiry (Tuesday)
    - If today is Wed/Thu/Fri → use NEXT week's expiry
    """
    this_week_expiry = get_weekly_expiry(today)

    if today.weekday() == 0:  # Monday
        return this_week_expiry
    else:
        # Wed/Thu/Fri → next week's Tuesday
        next_tuesday = today + datetime.timedelta(days=(8 - today.weekday()))
        # Adjust for holidays
        while next_tuesday in INDIA_HOLIDAYS_2026 or next_tuesday.weekday() >= 5:
            next_tuesday -= datetime.timedelta(days=1)
        return next_tuesday


def is_expiry_day(today: datetime.date) -> bool:
    """
    Returns True if today IS the expiry day (should NOT trade).
    Expiry = Tuesday, or previous working day if Tuesday is holiday.
    """
    return today == get_weekly_expiry(today)


def is_trading_day(today: datetime.date) -> bool:
    """
    Returns True if we should trade today.
    Trade on Mon/Wed/Thu/Fri that are market days AND not expiry day.
    """
    # Must be a weekday and not a holiday
    if today.weekday() >= 5 or today in INDIA_HOLIDAYS_2026:
        return False
    # Must NOT be expiry day
    if is_expiry_day(today):
        return False
    return True

# ── EMA Calculator ────────────────────────────────────────────────────────────
def calc_ema(prices: list, period: int) -> float:
    if len(prices) < period:
        return sum(prices) / len(prices)
    k   = 2 / (period + 1)
    val = sum(prices[:period]) / period
    for p in prices[period:]:
        val = p * k + val * (1 - k)
    return round(val, 2)

# ── Angel One Helpers ─────────────────────────────────────────────────────────
def login_angel() -> AngelOneClient | None:
    try:
        client = AngelOneClient()
        if client.login():
            log.info("✅ Angel One login successful")
            return client
        log.error("❌ Angel One login failed")
        return None
    except Exception as e:
        log.error(f"❌ Login error: {e}")
        return None

def get_nifty_price(client: AngelOneClient) -> tuple[float, float, list] | None:
    """Fetch current NIFTY price + historical closes for EMA."""
    import urllib.request
    # Fetch historical data for EMA from Yahoo Finance as fallback
    try:
        url = "https://query1.finance.yahoo.com/v8/finance/chart/%5ENSEI?interval=1d&range=30d"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())
        closes = [c for c in data["chart"]["result"][0]["indicators"]["quote"][0]["close"] if c]
        # Try Angel One for most current price
        ltp = client.get_ltp("NSE", "Nifty 50", "99926000")
        current = ltp if ltp else closes[-1]
        return round(current, 2), round(closes[-2], 2), closes
    except Exception as e:
        log.warning(f"Price fetch error: {e}")
        return None

def is_bullish(current: float, prev: float, closes: list) -> tuple[bool, str]:
    """Check if both EMA9>EMA21 and price>prev close."""
    ema9  = calc_ema(closes, 9)
    ema21 = calc_ema(closes, 21)
    price_bull = current > prev
    ema_bull   = ema9 > ema21
    reason = f"Price={'↑' if price_bull else '↓'}({current} vs {prev}) EMA9={'↑' if ema_bull else '↓'}({ema9} vs {ema21})"
    return price_bull and ema_bull, reason

def get_atm_strike(price: float, width: int) -> int:
    return int(math.floor(price / width) * width)

def build_option_symbol(strike: int, expiry: datetime.date, opt_type: str) -> str:
    day = expiry.strftime("%d").lstrip("0")
    mon = expiry.strftime("%b").upper()
    yr  = expiry.strftime("%y")
    return f"NIFTY{day}{mon}{yr}{strike}{opt_type}"

def get_next_tuesday(today: datetime.date) -> datetime.date:
    """Get this week's Tuesday (today if Tuesday, else next Tuesday)."""
    days = (1 - today.weekday()) % 7
    return today + datetime.timedelta(days=days if days else 0)

def search_scrip_with_retry(client: AngelOneClient, exchange: str, symbol: str, retries: int = 5):
    """Search for a scrip with rate-limit retry + backoff. Returns results or None."""
    for attempt in range(retries):
        try:
            return client.search_scrip(exchange, symbol)
        except Exception as e:
            if "rate" in str(e).lower() and attempt < retries - 1:
                wait = 2 ** (attempt + 1)  # 2s, 4s, 8s, 16s, 32s backoff
                log.warning(f"Rate limit hit searching {symbol}, retry {attempt+1}/{retries} in {wait}s...")
                time.sleep(wait)
                continue
            log.warning(f"Scrip search failed for {symbol}: {e}")
            return None
    return None


def fetch_option_ltp(client: AngelOneClient, symbol: str, retries: int = 5) -> float | None:
    """Search and fetch LTP for an option contract. Retries on rate-limit errors."""
    for attempt in range(retries):
        try:
            results = client.search_scrip("NFO", symbol)
            if not results:
                log.warning(f"No scrip found: {symbol}")
                return None
            time.sleep(0.5)  # small gap between search and ltp call
            token = results[0]["symboltoken"]
            ltp   = client.get_ltp("NFO", symbol, token)
            return ltp
        except Exception as e:
            if "rate" in str(e).lower() and attempt < retries - 1:
                wait = 2 ** (attempt + 1)  # 2s, 4s, 8s, 16s, 32s backoff
                log.warning(f"Rate limit hit for {symbol}, retry {attempt+1}/{retries} in {wait}s...")
                time.sleep(wait)
                continue
            log.warning(f"Option LTP fetch failed for {symbol}: {e}")
            return None
    return None

def place_order(client: AngelOneClient, symbol: str, token: str,
                qty: int, order_type: str, price: float = 0) -> str | None:
    """
    Place a real order via Angel One.
    order_type: 'BUY' or 'SELL'
    Returns order_id on success, None on failure.
    """
    import urllib.request
    body = {
        "variety":          "NORMAL",
        "tradingsymbol":    symbol,
        "symboltoken":      token,
        "transactiontype":  order_type,
        "exchange":         "NFO",
        "ordertype":        "LIMIT",
        "producttype":      "INTRADAY",
        "duration":         "DAY",
        "price":            str(round(price * 1.02, 1)),  # 2% buffer for limit order
        "squareoff":        "0",
        "stoploss":         "0",
        "quantity":         str(qty),
    }
    import urllib.error
    url = "https://apiconnect.angelone.in/rest/secure/angelbroking/order/v1/placeOrder"
    data = json.dumps(body).encode()
    headers = {
        "Content-Type":     "application/json",
        "Accept":           "application/json",
        "X-UserType":       "USER",
        "X-SourceID":       "WEB",
        "X-ClientLocalIP":  "127.0.0.1",
        "X-ClientPublicIP": "127.0.0.1",
        "X-MACAddress":     "00:00:00:00:00:00",
        "X-PrivateKey":     os.environ["ANGEL_API_KEY"],
        "Authorization":    f"Bearer {client.jwt_token}",
    }
    try:
        req = urllib.request.Request(url, data=data, method="POST", headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
        if result.get("status"):
            order_id = result["data"]["orderid"]
            log.info(f"✅ Order placed: {order_type} {symbol} x{qty} @ {price} → ID: {order_id}")
            return order_id
        else:
            log.error(f"❌ Order failed: {result.get('message')}")
            return None
    except Exception as e:
        log.error(f"❌ Order placement error: {e}")
        return None


def get_order_status(client: AngelOneClient, order_id: str) -> str | None:
    """
    Fetch the current status of an order by its order ID.
    Returns status string like 'open', 'complete', 'rejected', 'cancelled', etc.
    Returns None on fetch failure.
    """
    import urllib.request
    url = "https://apiconnect.angelone.in/rest/secure/angelbroking/order/v1/getOrderBook"
    headers = {
        "Content-Type":     "application/json",
        "Accept":           "application/json",
        "X-UserType":       "USER",
        "X-SourceID":       "WEB",
        "X-ClientLocalIP":  "127.0.0.1",
        "X-ClientPublicIP": "127.0.0.1",
        "X-MACAddress":     "00:00:00:00:00:00",
        "X-PrivateKey":     os.environ["ANGEL_API_KEY"],
        "Authorization":    f"Bearer {client.jwt_token}",
    }
    try:
        req = urllib.request.Request(url, method="GET", headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
        if not result.get("status"):
            log.warning(f"Order book fetch failed: {result.get('message')}")
            return None
        for order in result.get("data") or []:
            if order.get("orderid") == order_id:
                return order.get("status", "").lower()
        log.warning(f"Order {order_id} not found in order book")
        return None
    except Exception as e:
        log.warning(f"Order status fetch error for {order_id}: {e}")
        return None


def wait_for_order_fill(client: AngelOneClient, order_id: str,
                        max_wait_seconds: int = 60, poll_seconds: int = 3) -> bool:
    """
    Poll order status until it's 'complete' (filled), or until it's clearly
    rejected/cancelled, or until max_wait_seconds elapses.
    Returns True only if the order actually executed/completed.
    """
    elapsed = 0
    while elapsed < max_wait_seconds:
        status = get_order_status(client, order_id)
        if status is None:
            log.warning(f"  Could not determine status of order {order_id}, retrying...")
        elif status in ("complete", "executed"):
            log.info(f"  ✅ Order {order_id} CONFIRMED EXECUTED (status: {status})")
            return True
        elif status in ("rejected", "cancelled", "canceled"):
            log.error(f"  ❌ Order {order_id} did NOT execute (status: {status})")
            return False
        else:
            log.info(f"  ⏳ Order {order_id} status: {status} — waiting...")
        time.sleep(poll_seconds)
        elapsed += poll_seconds
    log.error(f"  ⏰ Timed out waiting for order {order_id} to fill after {max_wait_seconds}s")
    return False

# ── Excel Journal ─────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="7B0000")   # Dark red for live trades
PRF_FILL  = PatternFill("solid", start_color="C6EFCE")
SL_FILL   = PatternFill("solid", start_color="FFC7CE")
FLT_FILL  = PatternFill("solid", start_color="FFEB9C")
WHT_FILL  = PatternFill("solid", start_color="FFFFFF")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))

LIVE_COLS = [
    ("Date",              "date",           13),
    ("Instrument",        "instrument",     13),
    ("Strategy",          "strategy",       20),
    ("Mode",              "mode",           18),
    ("Trend",             "trend",          11),
    ("Trend Reason",      "trend_reason",   38),
    ("Expiry",            "expiry",         13),
    ("Entry Time",        "entry_time",     13),
    ("Exit Time",         "exit_time",      13),
    ("Spot at Entry",     "spot_entry",     15),
    ("Buy Strike",        "buy_strike",     15),
    ("Sell Strike",       "sell_strike",    15),
    ("Lots",              "lots",           8),
    ("Units",             "units",          8),
    ("Buy LTP Entry",     "buy_ltp_entry",  15),
    ("Sell LTP Entry",    "sell_ltp_entry", 16),
    ("Net Debit",         "net_debit",      13),
    ("Total Cost (Rs.)",  "total_cost",     15),
    ("Target (Rs.)",      "target",         13),
    ("Stop Loss (Rs.)",   "stop_loss",      14),
    ("Buy LTP Exit",      "buy_ltp_exit",   14),
    ("Sell LTP Exit",     "sell_ltp_exit",  15),
    ("Gross P&L (Rs.)",   "gross_pnl",      15),
    ("Brokerage (Rs.)",   "brokerage",      14),
    ("STT (Rs.)",         "stt",            11),
    ("SEBI Fee (Rs.)",    "sebi_fee",       13),
    ("Stamp Duty (Rs.)",  "stamp_duty",     15),
    ("GST (Rs.)",         "gst",            11),
    ("Total Fees (Rs.)",  "total_fees",     15),
    ("Net P&L (Rs.)",     "net_pnl",        14),
    ("Return %",          "return_pct",     12),
    ("Outcome",           "outcome",        12),
    ("Buy Order ID",      "buy_order_id",   18),
    ("Sell Order ID",     "sell_order_id",  18),
    ("Exit Reason",       "exit_reason",    20),
]

def _hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill      = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = THIN

def _data(ws, row, col, val, outcome=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = THIN
    c.fill = {"PROFIT": PRF_FILL, "STOP LOSS": SL_FILL,
              "FLAT": FLT_FILL}.get(outcome, WHT_FILL)

def ensure_live_sheet(wb):
    sheet_name = "NIFTY_LIVE_BullCall"
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)  # Insert as first sheet
    ws.merge_cells(f"A1:{get_column_letter(len(LIVE_COLS))}1")
    t       = ws["A1"]
    t.value = "NIFTY LIVE Trading — Bull Call Spread | REAL MONEY | Rs.15,000 Budget"
    t.font  = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    t.fill  = HDR_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for i, (h, _, w) in enumerate(LIVE_COLS, 1):
        _hdr(ws, 2, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 35
    ws.freeze_panes = "A3"
    return ws

def append_live_trade(trade: dict):
    """Load existing journal, append trade, save back — preserves all history."""
    if not os.path.exists(JOURNAL_FILE):
        log.error(f"Journal not found: {JOURNAL_FILE}")
        return
    wb  = openpyxl.load_workbook(JOURNAL_FILE)
    ws  = ensure_live_sheet(wb)
    row = ws.max_row + 1
    out = trade.get("outcome", "")
    for i, (_, key, _) in enumerate(LIVE_COLS, 1):
        _data(ws, row, i, trade.get(key, ""), out)
    wb.save(JOURNAL_FILE)
    log.info(f"✅ Trade logged to journal → row {row}")

# ── Fee Calculator ────────────────────────────────────────────────────────────
def calc_fees(lots: int, premium: float) -> dict:
    units     = lots * LOT_SIZE
    turnover  = premium * units
    brokerage = 20 * 2           # Rs.20 flat per order × 2 legs
    sebi      = 0.0001/100 * turnover * 2
    stt       = 0.0125/100 * turnover
    stamp     = 0.003/100 * turnover
    gst       = 0.18 * (brokerage + sebi)
    total     = round(brokerage + sebi + stt + stamp + gst, 2)
    return {
        "brokerage":  round(brokerage, 2),
        "stt":        round(stt, 2),
        "sebi_fee":   round(sebi, 2),
        "stamp_duty": round(stamp, 2),
        "gst":        round(gst, 2),
        "total_fees": total,
    }

# ── Main Trading Logic ────────────────────────────────────────────────────────
def run_live_trade() -> str:
    """
    Returns one of:
      'SKIPPED' — correctly skipped by design (holiday, expiry day, bearish trend)
      'ERROR'   — failed due to a technical error (login, price fetch, rate limit, etc.)
      'TRADED'  — trade was placed and completed (profit/loss/flat)
    """
    log.info("=" * 65)
    log.info("  NIFTY LIVE BOT — BULL CALL SPREAD — STARTING")
    log.info("=" * 65)

    today = datetime.date.today()
    if not is_trading_day(today):
        expiry_day = get_weekly_expiry(today)
        if is_expiry_day(today):
            log.info(f"{today} is expiry day ({expiry_day}) — skipping live trade.")
        else:
            log.info(f"{today} is not a market trading day — skipping.")
        return "SKIPPED"

    # Get correct expiry contract for today
    expiry = get_trade_expiry(today)
    log.info(f"  Today      : {today} ({today.strftime('%A')})")
    log.info(f"  Using expiry: {expiry} ({expiry.strftime('%A')})")

    # Login
    client = login_angel()
    if not client:
        log.error("Cannot proceed without Angel One connection.")
        return "ERROR"

    # Fetch price & check trend
    result = get_nifty_price(client)
    if not result:
        log.error("Cannot fetch NIFTY price. Aborting.")
        return "ERROR"

    current, prev, closes = result
    log.info(f"  NIFTY Spot  : {current:,.2f}")
    log.info(f"  Prev Close  : {prev:,.2f}")

    bullish, reason = is_bullish(current, prev, closes)
    log.info(f"  Trend       : {'BULLISH ✅' if bullish else 'BEARISH/SKIP ❌'}")
    log.info(f"  Reason      : {reason}")

    if not bullish:
        log.info("  Trend not bullish — skipping live trade today.")
        return "SKIPPED"

    atm        = get_atm_strike(current, STRIKE_WIDTH)
    buy_strike = atm
    sell_strike = atm + STRIKE_WIDTH

    buy_symbol  = build_option_symbol(buy_strike, expiry, "CE")
    sell_symbol = build_option_symbol(sell_strike, expiry, "CE")

    log.info(f"  Buy        : {buy_symbol}")
    log.info(f"  Sell       : {sell_symbol}")

    # Fetch live option premiums
    buy_ltp  = fetch_option_ltp(client, buy_symbol)
    time.sleep(1)   # Angel One rate limit — avoid back-to-back calls
    sell_ltp = fetch_option_ltp(client, sell_symbol)

    if not buy_ltp or not sell_ltp:
        log.error("Cannot fetch option premiums. Aborting.")
        return "ERROR"

    net_debit  = round(buy_ltp - sell_ltp, 2)
    if net_debit <= 0:
        log.error(f"Invalid net debit: {net_debit}. Aborting.")
        return "ERROR"

    # Calculate lots within budget.
    # IMPORTANT: Angel One requires you to first BUY the hedge leg, which costs the
    # FULL buy-side premium upfront (not just the net debit) — the margin benefit
    # on the SELL leg only applies AFTER the buy leg is confirmed filled. So lot
    # sizing must be based on what it costs to buy ONE leg, not the net spread cost.
    fees_est   = 40 + (0.0125/100 * buy_ltp * LOT_SIZE)
    cost_lot   = buy_ltp * LOT_SIZE + fees_est   # gross buy-leg cost per lot
    lots       = max(1, int(BUDGET / cost_lot))
    units      = lots * LOT_SIZE
    total_cost = round(net_debit * units, 2)       # net spread cost (for P&L/target/SL purposes)
    buy_cost   = round(buy_ltp * units, 2)          # what you actually need upfront for the BUY leg
    target     = round(total_cost * TARGET_PCT, 2)
    stop_loss  = round(total_cost * SL_PCT, 2)

    log.info(f"  Net Debit  : Rs.{net_debit}")
    log.info(f"  Buy Leg Cost (upfront, before hedge benefit): Rs.{buy_cost:,.2f}")
    log.info(f"  Lots       : {lots} ({units} units)")
    log.info(f"  Total Cost : Rs.{total_cost:,.2f}")
    log.info(f"  Target     : Rs.{target:,.2f} (+{TARGET_PCT*100:.0f}%)")
    log.info(f"  Stop Loss  : Rs.{stop_loss:,.2f} (-{SL_PCT*100:.0f}%)")

    # Place entry orders
    log.info("\n  Placing entry orders...")
    buy_results  = search_scrip_with_retry(client, "NFO", buy_symbol)
    time.sleep(1)   # Angel One rate limit — avoid back-to-back calls
    sell_results = search_scrip_with_retry(client, "NFO", sell_symbol)

    if not buy_results or not sell_results:
        log.error("Cannot find option tokens. Aborting.")
        return "ERROR"

    buy_token  = buy_results[0]["symboltoken"]
    sell_token = sell_results[0]["symboltoken"]

    # ── STEP 1: Place BUY leg first ───────────────────────────────────────
    log.info("  Placing BUY leg (hedge must fill before SELL leg)...")
    buy_order_id = place_order(client, buy_symbol, buy_token, units, "BUY", buy_ltp)

    if not buy_order_id:
        log.error("❌ BUY leg order placement failed. No position taken. Aborting safely.")
        return "ERROR"

    # ── STEP 2: Wait for BUY leg to actually be CONFIRMED EXECUTED ────────
    log.info(f"  Waiting for BUY order {buy_order_id} to fill before placing SELL leg...")
    buy_filled = wait_for_order_fill(client, buy_order_id, max_wait_seconds=60, poll_seconds=3)

    if not buy_filled:
        log.error(f"❌ BUY leg {buy_order_id} did NOT execute (rejected/cancelled/timed out).")
        log.error("  No SELL leg will be placed. Check Angel One dashboard to confirm no naked position exists.")
        return "ERROR"

    # ── STEP 3: BUY confirmed filled — NOW place SELL leg (gets margin hedge benefit) ──
    log.info("  ✅ BUY leg confirmed filled — placing SELL leg now (should get hedge margin benefit)...")
    sell_order_id = place_order(client, sell_symbol, sell_token, units, "SELL", sell_ltp)

    if not sell_order_id:
        log.error(f"⚠️  CRITICAL: BUY order {buy_order_id} is FILLED but SELL leg FAILED to place.")
        log.error("  You currently hold a NAKED LONG CALL position (not a spread). Check Angel One dashboard immediately!")
        log.error("  Consider manually squaring off the BUY leg if this was not intended.")
        return "ERROR"

    # Confirm SELL leg also actually filled (not just accepted)
    sell_filled = wait_for_order_fill(client, sell_order_id, max_wait_seconds=60, poll_seconds=3)
    if not sell_filled:
        log.error(f"⚠️  CRITICAL: BUY order {buy_order_id} is FILLED but SELL order {sell_order_id} did NOT execute.")
        log.error("  You currently hold a NAKED LONG CALL position (not a spread). Check Angel One dashboard immediately!")
        return "ERROR"

    entry_time = datetime.datetime.now(IST)
    log.info(f"  ✅ Both legs confirmed filled — spread entry complete at {entry_time.strftime('%H:%M IST')}")

    # Monitor position
    log.info(f"\n  Monitoring position every {POLL_MINS} minutes...")
    log.info(f"  Will force-exit at {EXIT_HOUR}:{EXIT_MIN:02d} IST if not hit target/SL")

    outcome     = "FLAT"
    exit_reason = "End of day exit"
    buy_exit    = buy_ltp
    sell_exit   = sell_ltp

    while True:
        now = datetime.datetime.now(IST)

        # Force exit at 3:15 PM
        if now.hour > EXIT_HOUR or (now.hour == EXIT_HOUR and now.minute >= EXIT_MIN):
            log.info("  ⏰ 3:15 PM reached — force closing position")
            exit_reason = "End of day force exit"
            outcome     = "FLAT"
            break

        time.sleep(POLL_MINS * 60)

        # Re-login if needed (token valid 1 day but refresh for safety)
        try:
            curr_buy  = fetch_option_ltp(client, buy_symbol)  or buy_ltp
            time.sleep(1)   # Angel One rate limit — avoid back-to-back calls
            curr_sell = fetch_option_ltp(client, sell_symbol) or sell_ltp
            curr_net  = curr_buy - curr_sell
            curr_pnl  = round((curr_net - net_debit) * units, 2)

            log.info(f"  [{now.strftime('%H:%M')}] Buy:{curr_buy} Sell:{curr_sell} Net:{curr_net:.2f} P&L:Rs.{curr_pnl:,.2f}")

            if curr_pnl >= target:
                outcome     = "PROFIT"
                exit_reason = "Target hit"
                buy_exit    = curr_buy
                sell_exit   = curr_sell
                log.info(f"  🎯 TARGET HIT! P&L: Rs.{curr_pnl:,.2f}")
                break
            elif curr_pnl <= -stop_loss:
                outcome     = "STOP LOSS"
                exit_reason = "Stop loss hit"
                buy_exit    = curr_buy
                sell_exit   = curr_sell
                log.info(f"  🛑 STOP LOSS HIT! P&L: Rs.{curr_pnl:,.2f}")
                break

        except Exception as e:
            log.warning(f"  Monitoring error: {e} — continuing")
            continue

    # Place exit orders (reverse)
    log.info("\n  Placing exit orders...")
    place_order(client, buy_symbol,  buy_token,  units, "SELL", buy_exit)
    place_order(client, sell_symbol, sell_token, units, "BUY",  sell_exit)

    exit_time  = datetime.datetime.now(IST)
    gross_pnl  = round((buy_exit - sell_exit - net_debit) * units, 2)
    fees       = calc_fees(lots, net_debit)
    net_pnl    = round(gross_pnl - fees["total_fees"], 2)
    return_pct = round(gross_pnl / total_cost * 100, 1) if total_cost else 0

    # Log to journal
    trade = {
        "date":           today.strftime("%Y-%m-%d"),
        "instrument":     "NIFTY",
        "strategy":       "Bull Call Spread",
        "mode":           "🔴 LIVE — REAL MONEY",
        "trend":          "BULLISH",
        "trend_reason":   reason,
        "expiry":         expiry.strftime("%Y-%m-%d"),
        "entry_time":     entry_time.strftime("%H:%M IST"),
        "exit_time":      exit_time.strftime("%H:%M IST"),
        "spot_entry":     current,
        "buy_strike":     f"{buy_strike} CE",
        "sell_strike":    f"{sell_strike} CE",
        "lots":           lots,
        "units":          units,
        "buy_ltp_entry":  buy_ltp,
        "sell_ltp_entry": sell_ltp,
        "net_debit":      net_debit,
        "total_cost":     total_cost,
        "target":         target,
        "stop_loss":      stop_loss,
        "buy_ltp_exit":   buy_exit,
        "sell_ltp_exit":  sell_exit,
        "gross_pnl":      gross_pnl,
        "brokerage":      fees["brokerage"],
        "stt":            fees["stt"],
        "sebi_fee":       fees["sebi_fee"],
        "stamp_duty":     fees["stamp_duty"],
        "gst":            fees["gst"],
        "total_fees":     fees["total_fees"],
        "net_pnl":        net_pnl,
        "return_pct":     return_pct,
        "outcome":        outcome,
        "buy_order_id":   buy_order_id,
        "sell_order_id":  sell_order_id,
        "exit_reason":    exit_reason,
    }

    append_live_trade(trade)

    log.info("\n" + "=" * 65)
    log.info(f"  LIVE TRADE COMPLETE")
    log.info(f"  Outcome    : {outcome}")
    log.info(f"  Gross P&L  : Rs.{gross_pnl:,.2f}")
    log.info(f"  Total Fees : Rs.{fees['total_fees']:,.2f}")
    log.info(f"  Net P&L    : Rs.{net_pnl:,.2f}")
    log.info(f"  Return %   : {return_pct}%")
    log.info("=" * 65)

    return "TRADED"


if __name__ == "__main__":
    status = run_live_trade()
    log.info(f"\nFinal status: {status}")
    # Exit code 0 for SKIPPED/TRADED (no retry needed), exit code 1 for ERROR (scheduler should retry)
    if status == "ERROR":
        sys.exit(1)
    sys.exit(0)
