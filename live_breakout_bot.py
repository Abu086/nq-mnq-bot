"""
Strategy Four — Live Breakout Bot (WebSocket Version)
=======================================================
Uses Angel One's SmartWebSocketV2 for real-time tick data.
No rate limits — one persistent connection streams all stock prices.

HOW IT WORKS:
1. At startup: login, load universe (20 stocks from cache), subscribe via WebSocket
2. WebSocket streams live ticks (LTP, OHLC, Volume) for all subscribed stocks
3. Every tick: update our in-memory 15-min candle for that stock
4. At each 15-min candle close: apply all 6 filters, check for breakout signal
5. Signal found: place MIS order via REST API
6. Monitor open positions via LTP ticks (no separate monitoring loop needed)
7. At 3:15 PM: force exit all positions

FILTERS (same as backtest):
1. Volume >= 1.5x 20-period average
2. Breakout magnitude >= 0.3%
3. Trend alignment (price above/below 20-MA)
4. Tight consolidation < 1% range
5. Candle body >= 0.3% of entry price
6. ROC > 0.5% in breakout direction

PARAMETERS:
- Budget: Rs.23,000 own per trade (5x MIS = Rs.1,15,000 effective)
- Max concurrent: 3 positions (Rs.84,000 total / Rs.23,000)
- SL: 0.5%, Target: 2.0%
- Universe: top 20 from ranking cache
- No trade before 10:00 AM IST
"""

import datetime
import json
import logging
import os
import sys
import threading
import time
from collections import defaultdict
from zoneinfo import ZoneInfo

import pandas as pd

from angel_one_client import AngelOneClient, _request

# ── Logging ───────────────────────────────────────────────────────────────
logging.Formatter.converter = lambda *args: datetime.datetime.now(ZoneInfo("Asia/Kolkata")).timetuple()
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s IST] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("logs/live_breakout.log"),
    ],
)
log = logging.getLogger(__name__)

IST = ZoneInfo("Asia/Kolkata")

# ── Strategy Parameters ───────────────────────────────────────────────────
BUDGET           = 23_000
MARGIN_MULTIPLE  = 5
DAILY_BUDGET     = 84_000
MAX_CONCURRENT   = int(DAILY_BUDGET // BUDGET)   # 3
LOOKBACK         = 5
SL_PCT           = 0.010
TARGET_PCT       = 0.020
VOL_MULTIPLE     = 1.5
BREAKOUT_MIN_PCT = 0.3
CONSOLIDATION_MAX= 1.0
CANDLE_BODY_MIN  = 0.003
ROC_MIN          = 0.5
CANDLE_MINUTES   = 15
FORCE_EXIT_TIME  = datetime.time(15, 15)
ENTRY_START_TIME = datetime.time(10, 0)
ENTRY_END_TIME   = datetime.time(14, 59)
TOKEN_CACHE      = "nifty500_token_cache.json"
RANKING_CACHE    = "top100_ranking_cache.json"
TOP_N            = 100   # Use top 20 for WebSocket (well within 1000 limit)

INDIA_HOLIDAYS_2026 = {
    datetime.date(2026,  1, 26), datetime.date(2026,  2, 26),
    datetime.date(2026,  3, 17), datetime.date(2026,  4,  2),
    datetime.date(2026,  4, 14), datetime.date(2026,  4, 30),
    datetime.date(2026,  8, 15), datetime.date(2026,  8, 27),
    datetime.date(2026, 10,  2), datetime.date(2026, 10, 20),
    datetime.date(2026, 11,  9), datetime.date(2026, 11, 10),
    datetime.date(2026, 11, 25), datetime.date(2026, 12, 25),
}

def is_trading_day(dt: datetime.date) -> bool:
    return dt.weekday() < 5 and dt not in INDIA_HOLIDAYS_2026

# ── Order Placement (REST API) ────────────────────────────────────────────
_margin_cache = {"value": 0.0, "ts": 0}

def _get_cached_margin(client):
    now_ts = time.time()
    if now_ts - _margin_cache["ts"] > 30:
        _margin_cache["value"] = client.get_available_margin()
        _margin_cache["ts"] = now_ts
    return _margin_cache["value"]

def place_order(client, symbol, token, qty, side, price=0):
    import urllib.request
    # Use LIMIT order at LTP price — works for cautionary listing stocks too
    # MARKET orders are rejected for surveillance stocks (Angel One AB4036 error)
    limit_price = round(price if price else 0, 1)
    body = {
        "variety": "NORMAL", "tradingsymbol": symbol + "-EQ", "symboltoken": token,
        "transactiontype": side, "exchange": "NSE",
        "ordertype": "LIMIT" if limit_price > 0 else "MARKET",
        "producttype": "INTRADAY", "duration": "DAY",
        "price": str(limit_price) if limit_price > 0 else "0",
        "squareoff": "0", "stoploss": "0", "quantity": str(qty),
    }
    headers = {
        "Content-Type": "application/json", "Accept": "application/json",
        "X-UserType": "USER", "X-SourceID": "WEB",
        "X-ClientLocalIP": "127.0.0.1", "X-ClientPublicIP": "127.0.0.1",
        "X-MACAddress": "00:00:00:00:00:00",
        "X-PrivateKey": os.environ["ANGEL_API_KEY"],
        "Authorization": f"Bearer {client.jwt_token}",
    }
    log.info(f"ORDER BODY: {body}")
    try:
        url = "https://apiconnect.angelone.in/rest/secure/angelbroking/order/v1/placeOrder"
        req = urllib.request.Request(url, data=json.dumps(body).encode(),
                                     method="POST", headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
        if result.get("status"):
            oid = result["data"]["orderid"]
            log.info(f"✅ {side} {symbol} x{qty} → Order {oid}")
            return oid
        msg = result.get("message", "")
        if "mismatch" in msg.lower():
            import time as _t
            _t.sleep(2)
            with urllib.request.urlopen(req, timeout=15) as resp2:
                result2 = json.loads(resp2.read())
            if result2.get("status"):
                oid = result2["data"]["orderid"]
                log.info(f"✅ (retry) {side} {symbol} x{qty} → Order {oid}")
                return oid
            log.error(f"Order failed after retry: {result2.get('message')}")
            return None
        log.error(f"❌ Order failed: {msg}")
        return None
    except Exception as e:
        log.error(f"❌ Order error: {e}")
        return None

# ── Journal Logging ───────────────────────────────────────────────────────
def log_trade(symbol, direction, entry_price, exit_price, qty,
              entry_time, exit_time, exit_reason):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        JOURNAL = "trading_journal.xlsx"
        if not os.path.exists(JOURNAL):
            return
        wb = openpyxl.load_workbook(JOURNAL)
        sheet = "Strategy4_WS_Live"
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
            headers = ["Date","Symbol","Direction","Entry Time","Exit Time",
                      "Entry Price","Exit Price","Qty","Gross P&L","Brokerage",
                      "Net P&L","Exit Reason"]
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", start_color="1F4E79")
        else:
            ws = wb[sheet]
        if direction == "LONG":
            gross = round((exit_price - entry_price) * qty, 2)
        else:
            gross = round((entry_price - exit_price) * qty, 2)
        brok = 40
        net  = round(gross - brok, 2)
        row  = ws.max_row + 1
        data = [entry_time.strftime("%Y-%m-%d"), symbol, direction,
                entry_time.strftime("%H:%M IST"), exit_time.strftime("%H:%M IST"),
                entry_price, exit_price, qty, gross, brok, net, exit_reason]
        fill = PatternFill("solid", start_color="C6EFCE" if net > 0 else "FFC7CE")
        for i, v in enumerate(data, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.alignment = Alignment(horizontal="center")
            if i == 11: c.fill = fill
        wb.save(JOURNAL)
        log.info(f"✅ Logged: {symbol} {direction} Net Rs.{net:,.2f}")
    except Exception as e:
        log.warning(f"Journal log failed: {e}")

# ── Candle Builder ────────────────────────────────────────────────────────
class CandleBuilder:
    """
    Builds 15-min OHLC candles from WebSocket tick data in memory.
    Each tick contains LTP, volume, and OHLC of the day from Angel One.
    We aggregate these into 15-min bars ourselves.
    """
    def __init__(self):
        # {symbol: [list of completed candles]}
        self.candles = defaultdict(list)
        # {symbol: current open candle}
        self.current = {}
        self.lock = threading.Lock()

    def _candle_start(self, ts: datetime.datetime) -> datetime.datetime:
        """Round timestamp down to nearest 15-min boundary."""
        minute = (ts.minute // CANDLE_MINUTES) * CANDLE_MINUTES
        return ts.replace(minute=minute, second=0, microsecond=0)

    def on_tick(self, symbol: str, ltp: float, volume: int, ts: datetime.datetime):
        """Process a new tick and update candles."""
        with self.lock:
            bar_start = self._candle_start(ts)

            if symbol not in self.current:
                # First tick for this symbol
                self.current[symbol] = {
                    "open": ltp, "high": ltp, "low": ltp,
                    "close": ltp, "volume": volume, "start": bar_start
                }
                return

            curr = self.current[symbol]

            if bar_start > curr["start"]:
                # New 15-min bar started — save the completed candle
                self.candles[symbol].append({
                    "timestamp": curr["start"],
                    "Open":   curr["open"],
                    "High":   curr["high"],
                    "Low":    curr["low"],
                    "Close":  curr["close"],
                    "Volume": curr["volume"],
                })
                # Keep only last 100 candles per stock
                if len(self.candles[symbol]) > 100:
                    self.candles[symbol] = self.candles[symbol][-100:]
                # Start new candle
                self.current[symbol] = {
                    "open": ltp, "high": ltp, "low": ltp,
                    "close": ltp, "volume": volume, "start": bar_start
                }
            else:
                # Same bar — update OHLC
                curr["high"]   = max(curr["high"], ltp)
                curr["low"]    = min(curr["low"], ltp)
                curr["close"]  = ltp
                curr["volume"] = max(curr["volume"], volume)

    def get_dataframe(self, symbol: str) -> pd.DataFrame | None:
        """Return completed candles as a DataFrame."""
        with self.lock:
            candles = self.candles.get(symbol, [])
            if len(candles) < LOOKBACK + 2:
                return None
            df = pd.DataFrame(candles)
            df = df.set_index("timestamp")
            return df

    def last_completed_bar_time(self, symbol: str) -> datetime.datetime | None:
        """Return the timestamp of the most recently completed bar."""
        with self.lock:
            candles = self.candles.get(symbol, [])
            if not candles:
                return None
            return candles[-1]["timestamp"]

# ── Signal Detection ──────────────────────────────────────────────────────
def check_signal(df: pd.DataFrame, entry_price_hint: float) -> tuple | None:
    """Apply all 6 filters. Returns (direction, entry_price) or None."""
    if len(df) < LOOKBACK + 2:
        return None

    df = df.copy()
    df["prior_high"]      = df["High"].rolling(LOOKBACK).max().shift(1)
    df["prior_low"]       = df["Low"].rolling(LOOKBACK).min().shift(1)
    df["vol_avg20"]       = df["Volume"].rolling(20).mean().shift(1)
    df["ma20"]            = df["Close"].rolling(20).mean().shift(1)
    df["range_width_pct"] = (
        (df["High"].rolling(LOOKBACK).max() - df["Low"].rolling(LOOKBACK).min()) /
        df["prior_high"] * 100
    ).shift(1)

    row = df.iloc[-1]

    if pd.isna(row["prior_high"]) or pd.isna(row["vol_avg20"]) or pd.isna(row["ma20"]):
        return None

    # Breakout detection
    if row["High"] > row["prior_high"]:
        direction, entry_price = "LONG", row["prior_high"]
    elif row["Low"] < row["prior_low"]:
        direction, entry_price = "SHORT", row["prior_low"]
    else:
        return None

    # Filter 1: Volume
    if row["Volume"] < VOL_MULTIPLE * row["vol_avg20"]:
        return None
    # Filter 2: Breakout magnitude
    if direction == "LONG":
        mag = (row["High"] - row["prior_high"]) / row["prior_high"] * 100
    else:
        mag = (row["prior_low"] - row["Low"]) / row["prior_low"] * 100
    if mag < BREAKOUT_MIN_PCT:
        return None
    # Filter 3: Trend alignment
    if direction == "LONG" and row["Close"] < row["ma20"]:
        return None
    if direction == "SHORT" and row["Close"] > row["ma20"]:
        return None
    # Filter 4: Tight consolidation
    if pd.isna(row["range_width_pct"]) or row["range_width_pct"] > CONSOLIDATION_MAX:
        return None
    # Filter 5: Candle body
    if abs(row["Close"] - row["Open"]) < entry_price * CANDLE_BODY_MIN:
        return None
    # Filter 6: ROC
    if len(df) >= 6:
        roc = (df["Close"].iloc[-1] - df["Close"].iloc[-6]) / df["Close"].iloc[-6] * 100
        if direction == "LONG" and roc < ROC_MIN:
            return None
        if direction == "SHORT" and roc > -ROC_MIN:
            return None

    return direction, round(entry_price, 2)

# ── Main Bot ──────────────────────────────────────────────────────────────
def run():
    log.info("=" * 65)
    log.info("  STRATEGY FOUR LIVE BOT — WEBSOCKET VERSION")
    log.info(f"  Per trade: Rs.{BUDGET:,} x{MARGIN_MULTIPLE} margin | Max positions: unlimited (live margin check)")
    log.info("=" * 65)

    today = datetime.date.today()
    if not is_trading_day(today):
        log.info(f"{today} is not a trading day.")
        return "SKIPPED"

    # Login
    client = AngelOneClient()
    if not client.login():
        log.error("Login failed.")
        return "ERROR"
    log.info("✅ Angel One login successful")

    # Load universe from cache
    if not os.path.exists(RANKING_CACHE):
        log.error("Ranking cache not found. Run update_universe.py first.")
        return "ERROR"
    with open(RANKING_CACHE) as f:
        cache = json.load(f)
    symbols = cache.get("symbols", [])[:TOP_N]
    if not symbols:
        log.error("No symbols in cache.")
        return "ERROR"
    log.info(f"Universe: {len(symbols)} stocks")

    # Load tokens
    with open(TOKEN_CACHE) as f:
        token_map = json.load(f)
    missing = [s for s in symbols if s not in token_map]
    if missing:
        log.warning(f"Missing tokens for: {missing}")
        symbols = [s for s in symbols if s in token_map]

    # State
    candle_builder  = CandleBuilder()
    open_positions  = {}    # symbol → position dict
    closed_positions = []
    traded_today    = set()
    daily_pnl       = 0.0
    last_checked    = {}    # symbol → last bar timestamp we checked signals for
    running         = [True]

    # DIAGNOSTIC (added 2026-07-14): zero trades on 2026-07-13 and 07-14
    # with no errors logged. Leading suspect is the `isinstance(message,
    # bytes): return` branch in on_data silently dropping every tick if
    # the installed SmartWebSocketV2 SDK delivers raw bytes instead of a
    # pre-parsed dict -- that would explain total silence with no errors.
    # This counter block only adds visibility, it changes no trading
    # behavior. Remove once the real cause is confirmed and fixed.
    _tick_diag = {
        "logged_first": False, "bytes_count": 0, "dict_count": 0,
        "other_count": 0, "no_symbol_count": 0, "no_ltp_count": 0,
    }


    # ── Pre-load today's candles via Yahoo Finance ────────────────────────
    log.info("Pre-loading today candles via Yahoo Finance...")
    import urllib.request as _ur
    import datetime as _dt
    _preloaded = 0
    for _sym in symbols:
        try:
            _url = f"https://query1.finance.yahoo.com/v8/finance/chart/{_sym}.NS?interval=15m&range=1d"
            _req = _ur.Request(_url, headers={"User-Agent": "Mozilla/5.0"})
            with _ur.urlopen(_req, timeout=10) as _r:
                _data = json.loads(_r.read())
            _res = _data["chart"]["result"][0]
            _ts = _res["timestamp"]
            _q = _res["indicators"]["quote"][0]
            for _j, _t in enumerate(_ts):
                if _q["open"][_j] and _q["high"][_j] and _q["low"][_j] and _q["close"][_j]:
                    _dt2 = _dt.datetime.fromtimestamp(_t, tz=IST)
                    with candle_builder.lock:
                        candle_builder.candles[_sym].append({"timestamp":_dt2,"Open":_q["open"][_j],"High":_q["high"][_j],"Low":_q["low"][_j],"Close":_q["close"][_j],"Volume":_q["volume"][_j] or 0})
            _preloaded += 1
        except:
            pass
        time.sleep(0.3)
    log.info(f"Pre-loaded {_preloaded}/{len(symbols)} stocks — signals fire immediately!")
    # ── End pre-load ──────────────────────────────────────────────────────

    # ── WebSocket Setup ───────────────────────────────────────────────────
    from SmartApi.smartWebSocketV2 import SmartWebSocketV2

    AUTH_TOKEN  = client.jwt_token
    FEED_TOKEN  = client.feed_token
    CLIENT_CODE = os.environ["ANGEL_CLIENT_ID"]

    sws = SmartWebSocketV2(
        auth_token=AUTH_TOKEN,
        api_key=os.environ["ANGEL_API_KEY"],
        client_code=CLIENT_CODE,
        feed_token=FEED_TOKEN,
    )

    # Subscribe to QUOTE mode (gives LTP + OHLC + Volume per tick)
    token_list = [{
        "exchangeType": 1,   # NSE_CM = 1
        "tokens": [token_map[s] for s in symbols]
    }]

    # Reverse map: token → symbol
    token_to_symbol = {token_map[s]: s for s in symbols}

    def on_data(wsapp, message):
        """Called for every tick received from WebSocket."""
        try:
            now = datetime.datetime.now(IST)

            if not _tick_diag["logged_first"]:
                _tick_diag["logged_first"] = True
                log.info(f"DIAG: first WS message type={type(message).__name__} "
                         f"sample={str(message)[:300]}")

            # DIAGNOSTIC guard (added 2026-07-14): count anything that's
            # neither bytes nor dict, WITHOUT disturbing the existing
            # bytes/dict branches or the large block nested inside the
            # dict branch below (kept at its original indentation on
            # purpose -- restructuring that block risked a much riskier
            # edit in live order-placement code).
            if not isinstance(message, (bytes, dict)):
                _tick_diag["other_count"] += 1
                return

            # Parse tick data
            if isinstance(message, bytes):
                # Binary format from SmartWebSocketV2
                # The SDK parses this automatically in newer versions
                _tick_diag["bytes_count"] += 1
                return
            if isinstance(message, dict):
                _tick_diag["dict_count"] += 1
                token      = str(message.get("token", ""))
                ltp        = message.get("last_traded_price", 0) / 100  # paise to rupees
                volume     = message.get("volume_trade_for_the_day", 0)
                symbol     = token_to_symbol.get(token)
                if not symbol:
                    _tick_diag["no_symbol_count"] += 1
                    return
                if ltp <= 0:
                    _tick_diag["no_ltp_count"] += 1
                    return

                # Update candle builder
                candle_builder.on_tick(symbol, ltp, volume, now)

                # Check if we're in a position — monitor SL/target
                if symbol in open_positions:
                    pos = open_positions[symbol]
                    direction = pos["direction"]
                    hit_target = hit_stop = False

                    if direction == "LONG":
                        hit_target = ltp >= pos["target"]
                        hit_stop   = ltp <= pos["stop"]
                    else:
                        hit_target = ltp <= pos["target"]
                        hit_stop   = ltp >= pos["stop"]

                    exit_reason = None
                    if hit_target:
                        exit_reason = "TARGET"
                    elif hit_stop:
                        exit_reason = "STOP LOSS"
                    elif now.time() >= FORCE_EXIT_TIME:
                        exit_reason = "FORCE EXIT (EOD)"

                    if exit_reason:
                        side = "SELL" if direction == "LONG" else "BUY"
                        place_order(client, symbol, token_map[symbol], pos["qty"], side, price=ltp)
                        exit_price = pos["target"] if "TARGET" in exit_reason else (
                            pos["stop"] if "STOP" in exit_reason else ltp)
                        if direction == "LONG":
                            pnl = round((exit_price - pos["entry_price"]) * pos["qty"], 2)
                        else:
                            pnl = round((pos["entry_price"] - exit_price) * pos["qty"], 2)
                        daily_pnl_ref[0] += pnl
                        icon = "🎯" if "TARGET" in exit_reason else "🛑"
                        log.info(f"  {icon} {symbol} {exit_reason} | P&L: Rs.{pnl:,.2f}")
                        log_trade(symbol, direction, pos["entry_price"], exit_price,
                                 pos["qty"], pos["entry_time"], now, exit_reason)
                        del open_positions[symbol]
                        closed_positions.append({**pos, "exit_price": exit_price,
                                                 "pnl": pnl, "exit_reason": exit_reason})
                    return

                # Check for new signals (only after 10 AM, only if slot available)
                if (now.time() < ENTRY_START_TIME or
                        now.time() >= ENTRY_END_TIME or
                        symbol in traded_today or
                        _get_cached_margin(client) < BUDGET):
                    return

                # Only check at 15-min candle boundaries
                last_bar = candle_builder.last_completed_bar_time(symbol)
                if last_bar is None or last_bar == last_checked.get(symbol):
                    return
                last_checked[symbol] = last_bar

                # Get candle data and check signal
                df = candle_builder.get_dataframe(symbol)
                if df is None:
                    return

                signal = check_signal(df, ltp)
                if signal is None:
                    return

                direction, _sig_price = signal
                entry_price = ltp
                qty = max(1, int((BUDGET * MARGIN_MULTIPLE) // ltp))

                if direction == "LONG":
                    stop   = round(entry_price * (1 - SL_PCT), 2)
                    target = round(entry_price * (1 + TARGET_PCT), 2)
                    side   = "BUY"
                else:
                    stop   = round(entry_price * (1 + SL_PCT), 2)
                    target = round(entry_price * (1 - TARGET_PCT), 2)
                    side   = "SELL"

                log.info(f"  🚀 SIGNAL: {symbol} {direction} @ {entry_price} "
                         f"| SL:{stop} | T:{target} | Qty:{qty}")

                order_id = place_order(client, symbol, token_map[symbol], qty, side, price=ltp)
                if order_id:
                    open_positions[symbol] = {
                        "symbol": symbol, "direction": direction,
                        "entry_price": entry_price, "qty": qty,
                        "stop": stop, "target": target,
                        "order_id": order_id, "entry_time": now,
                    }
                    traded_today.add(symbol)

        except Exception as e:
            log.warning(f"Tick processing error: {e}")

    def on_open(wsapp):
        log.info("✅ WebSocket connected — subscribing to stocks...")
        sws.subscribe("strategy4", 2, token_list)  # mode 2 = QUOTE
        log.info(f"✅ Subscribed to {len(symbols)} stocks")
        log.info(f"  Monitoring | Force exit at {FORCE_EXIT_TIME} IST")

    def on_error(wsapp, *args):
        # BUGFIX (2026-07-14): was `log.error(f"WebSocket error: {error}")`
        # -- referenced an undefined variable `error`, so any actual
        # websocket error raised a NameError instead of logging it, which
        # crashed the process (confirmed in last night's 2026-07-14 02:00
        # IST log: "tearing down on exception name 'error' is not defined").
        log.error(f"WebSocket error: {args}")

    def on_close(wsapp, *args):
        log.info("WebSocket connection closed")

    # Shared mutable P&L reference
    daily_pnl_ref = [0.0]

    sws.on_open    = on_open
    sws.on_error   = on_error
    sws.on_close   = on_close

    # Bind on_data as instance method

    sws.on_data = on_data

    # ── Force exit thread ─────────────────────────────────────────────────
    def force_exit_thread():
        while running[0]:
            now = datetime.datetime.now(IST)
            if now.time() >= FORCE_EXIT_TIME:
                for sym, pos in list(open_positions.items()):
                    side = "SELL" if pos["direction"] == "LONG" else "BUY"
                    place_order(client, sym, token_map[sym], pos["qty"], side)
                    log.info(f"  ⏰ Force exit: {sym}")
                running[0] = False
                sws.close_connection()
                break
            time.sleep(30)

    exit_thread = threading.Thread(target=force_exit_thread, daemon=True)
    exit_thread.start()

    # -- Token refresh thread (Angel One JWT expires after a few hours) ----
    def token_refresh_thread():
        while running[0]:
            time.sleep(3600)  # every 60 minutes
            try:
                if client.login():
                    log.info("🔄 Session token refreshed")
                else:
                    log.error("Token refresh failed")
            except Exception as e:
                log.error(f"Token refresh error: {e}")

    refresh_thread = threading.Thread(target=token_refresh_thread, daemon=True)
    refresh_thread.start()

    # ── Status thread ─────────────────────────────────────────────────────
    def status_thread():
        while running[0]:
            time.sleep(300)  # Every 5 minutes
            now = datetime.datetime.now(IST)
            log.info(f"  [{now.strftime('%H:%M')}] Open: {len(open_positions)} "
                     f"| Closed: {len(closed_positions)} | P&L: Rs.{daily_pnl_ref[0]:,.2f}")
            log.info(f"  DIAG ticks: bytes={_tick_diag['bytes_count']} "
                     f"dict={_tick_diag['dict_count']} other={_tick_diag['other_count']} "
                     f"no_symbol={_tick_diag['no_symbol_count']} no_ltp={_tick_diag['no_ltp_count']}")

    st = threading.Thread(target=status_thread, daemon=True)
    st.start()

    # ── Start WebSocket (blocking) ─────────────────────────────────────────
    log.info("Starting WebSocket connection...")
    sws.connect()

    # ── End of day summary ────────────────────────────────────────────────
    log.info("\n" + "=" * 65)
    log.info(f"  END OF DAY | Trades: {len(closed_positions)} | P&L: Rs.{daily_pnl_ref[0]:,.2f}")
    brok = 40 * len(closed_positions)
    log.info(f"  Brokerage: Rs.{brok:,} | Net: Rs.{daily_pnl_ref[0]-brok:,.2f}")
    log.info("=" * 65)
    return "TRADED" if closed_positions else "SKIPPED"


if __name__ == "__main__":
    status = run()
    sys.exit(1 if status == "ERROR" else 0)
