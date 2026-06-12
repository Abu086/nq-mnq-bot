"""
NQ / MNQ Bull Call Spread — Paper Trading Bot
Runs daily at 10:00 ET via GitHub Actions.
Data source : Yahoo Finance (free, no API key needed)
Journal     : trading_journal.xlsx (auto-updated)
"""

import datetime
import json
import logging
import math
import os
import random
import sys
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side)
from openpyxl.utils import get_column_letter

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s IST] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("logs/bot.log"),
    ],
)
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
ET = ZoneInfo("America/New_York")
JOURNAL_FILE = "trading_journal.xlsx"

INSTRUMENTS = {
    "NQ": {
        "ticker":           "NQ=F",
        "multiplier":       20,          # $20 per index point
        "strike_width":     25,          # points between strikes
        "contracts":        1,
        "budget":           1000,
        "cme_fee":          1.18,        # per leg per contract
        "broker_fee":       2.50,
        "nfa_fee":          0.02,
    },
    "MNQ": {
        "ticker":           "MNQ=F",
        "multiplier":       2,           # $2 per index point
        "strike_width":     25,
        "contracts":        3,           # 3 spreads within $1 000
        "budget":           1000,
        "cme_fee":          0.30,
        "broker_fee":       0.50,
        "nfa_fee":          0.02,
    },
}

TARGET_PCT   = 0.30   # exit at 30 % of max profit
SL_PCT       = 0.30   # stop loss at 30 % of premium paid


# ── Fee helpers ───────────────────────────────────────────────────────────────
def calc_fees(inst: dict, n_contracts: int) -> dict:
    """Return per-leg, round-trip, and total fees for a spread."""
    per_leg   = inst["cme_fee"] + inst["broker_fee"] + inst["nfa_fee"]
    legs      = 2          # buy leg + sell leg
    rt        = per_leg * legs * n_contracts
    return {
        "cme_fee_per_leg":    round(inst["cme_fee"]    * n_contracts, 4),
        "broker_fee_per_leg": round(inst["broker_fee"] * n_contracts, 4),
        "nfa_fee_per_leg":    round(inst["nfa_fee"]    * n_contracts, 4),
        "total_per_leg":      round(per_leg            * n_contracts, 4),
        "round_trip_total":   round(rt, 4),
    }


# ── Price fetcher (Yahoo Finance via urllib) ──────────────────────────────────
def fetch_price(ticker: str) -> float | None:
    """Fetch latest price for a futures ticker from Yahoo Finance."""
    import urllib.request, urllib.error
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
        f"?interval=1m&range=1d"
    )
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        req  = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
        closes = data["chart"]["result"][0]["indicators"]["quote"][0]["close"]
        closes = [c for c in closes if c is not None]
        return round(closes[-1], 2) if closes else None
    except Exception as exc:
        log.warning(f"Price fetch failed for {ticker}: {exc}")
        return None


def get_atm_strike(price: float, width: int) -> int:
    """Round price down to nearest strike width multiple."""
    return int(math.floor(price / width) * width)


# ── Simulate intraday P&L ─────────────────────────────────────────────────────
def simulate_trade(inst_name: str, inst: dict, spot: float) -> dict:
    """
    Simulate a Bull Call Spread entry at 10:00 ET and exit during the day.
    Premium is estimated as ~0.3 % of strike width * multiplier for ATM spreads.
    """
    width       = inst["strike_width"]
    multiplier  = inst["multiplier"]
    n           = inst["contracts"]

    buy_strike  = get_atm_strike(spot, width)
    sell_strike = buy_strike + width

    # Rough ATM spread premium: ~35–55 % of max value (randomised for simulation)
    max_value       = width * multiplier          # e.g. 25 × $20 = $500 for NQ
    entry_pct       = random.uniform(0.35, 0.55)
    entry_premium   = round(max_value * entry_pct, 2)   # per contract
    total_cost      = round(entry_premium * n, 2)

    target_profit   = round(max_value * TARGET_PCT * n, 2)
    stop_loss       = round(entry_premium * SL_PCT  * n, 2)

    # Simulate outcome: 55 % win, 30 % hit SL, 15 % flat
    rand = random.random()
    if rand < 0.55:
        outcome      = "PROFIT"
        exit_premium = round(entry_premium + (max_value - entry_premium) * random.uniform(0.2, 0.6), 2)
    elif rand < 0.85:
        outcome      = "STOP LOSS"
        exit_premium = round(entry_premium * (1 - SL_PCT) * random.uniform(0.5, 0.9), 2)
    else:
        outcome      = "FLAT"
        exit_premium = round(entry_premium * random.uniform(0.95, 1.05), 2)

    gross_pnl = round((exit_premium - entry_premium) * n, 2)
    fees      = calc_fees(inst, n)
    net_pnl   = round(gross_pnl - fees["round_trip_total"], 2)

    now_et    = datetime.datetime.now(ET)
    entry_t   = now_et.replace(hour=10, minute=0, second=0, microsecond=0)
    exit_mins = random.randint(90, 360)
    exit_t    = entry_t + datetime.timedelta(minutes=exit_mins)

    return {
        "date":                 now_et.strftime("%Y-%m-%d"),
        "instrument":           inst_name,
        "strategy":             "Bull Call Spread",
        "entry_time":           entry_t.strftime("%H:%M ET"),
        "exit_time":            exit_t.strftime("%H:%M ET"),
        "spot_at_entry":        spot,
        "buy_strike":           buy_strike,
        "sell_strike":          sell_strike,
        "contracts":            n,
        "entry_premium":        entry_premium,
        "exit_premium":         exit_premium,
        "total_cost":           total_cost,
        "max_value":            max_value * n,
        "target_profit":        target_profit,
        "stop_loss_amt":        stop_loss,
        "gross_pnl":            gross_pnl,
        "cme_fee":              fees["cme_fee_per_leg"] * 2,
        "broker_fee":           fees["broker_fee_per_leg"] * 2,
        "nfa_fee":              fees["nfa_fee_per_leg"] * 2,
        "total_fees":           fees["round_trip_total"],
        "net_pnl":              net_pnl,
        "outcome":              outcome,
        "budget":               inst["budget"],
        "mode":                 "PAPER TRADING",
    }


# ── Excel journal ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
PROFIT_FILL   = PatternFill("solid", start_color="C6EFCE")
LOSS_FILL     = PatternFill("solid", start_color="FFC7CE")
FLAT_FILL     = PatternFill("solid", start_color="FFEB9C")
SUBHDR_FILL   = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
THIN          = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

TRADE_COLS = [
    ("Date",            "date",           14),
    ("Instrument",      "instrument",     12),
    ("Strategy",        "strategy",       18),
    ("Entry Time",      "entry_time",     12),
    ("Exit Time",       "exit_time",      12),
    ("Spot at Entry",   "spot_at_entry",  14),
    ("Buy Strike",      "buy_strike",     12),
    ("Sell Strike",     "sell_strike",    12),
    ("Contracts",       "contracts",       11),
    ("Entry Premium",   "entry_premium",  15),
    ("Exit Premium",    "exit_premium",   14),
    ("Total Cost ($)",  "total_cost",     14),
    ("Max Value ($)",   "max_value",      13),
    ("Target ($)",      "target_profit",  12),
    ("Stop Loss ($)",   "stop_loss_amt",  13),
    ("Gross P&L ($)",   "gross_pnl",      13),
    ("CME Fee ($)",     "cme_fee",        12),
    ("Broker Fee ($)",  "broker_fee",     13),
    ("NFA Fee ($)",     "nfa_fee",        11),
    ("Total Fees ($)",  "total_fees",     13),
    ("Net P&L ($)",     "net_pnl",        13),
    ("Outcome",         "outcome",        12),
    ("Budget ($)",      "budget",         12),
    ("Mode",            "mode",           22),
]


def _hdr_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill      = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = THIN
    return c


def _data_cell(ws, row, col, value, outcome=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = THIN
    if outcome == "PROFIT":
        c.fill = PROFIT_FILL
    elif outcome == "STOP LOSS":
        c.fill = LOSS_FILL
    elif outcome == "FLAT":
        c.fill = FLAT_FILL
    else:
        c.fill = WHITE_FILL
    return c


def ensure_journal():
    """Create journal with all sheets if it doesn't exist."""
    if os.path.exists(JOURNAL_FILE):
        return openpyxl.load_workbook(JOURNAL_FILE)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheets = ["NQ_Trades", "MNQ_Trades", "US_Stocks", "Fee_Breakdown", "Summary"]
    for name in sheets:
        ws = wb.create_sheet(name)
        if name in ("NQ_Trades", "MNQ_Trades"):
            _build_trade_sheet(ws, name.split("_")[0])
        elif name == "US_Stocks":
            _build_stocks_sheet(ws)
        elif name == "Fee_Breakdown":
            _build_fee_sheet(ws)
        elif name == "Summary":
            _build_summary_sheet(ws)

    wb.save(JOURNAL_FILE)
    return wb


def _build_trade_sheet(ws, label):
    ws.row_dimensions[1].height = 35
    for i, (hdr, _, width) in enumerate(TRADE_COLS, 1):
        _hdr_cell(ws, 1, i, hdr)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    # Title banner
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(len(TRADE_COLS))}1")
    title = ws["A1"]
    title.value     = f"{label} Bull Call Spread — Paper Trading Journal"
    title.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    title.fill      = PatternFill("solid", start_color="0D2137")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"


def _build_stocks_sheet(ws):
    headers = [
        ("Date", 12), ("Ticker", 10), ("Company", 22), ("Action", 10),
        ("Qty", 8), ("Entry Price", 13), ("Exit Price", 12),
        ("Gross P&L ($)", 14), ("Broker Fee ($)", 14), ("Net P&L ($)", 13),
        ("Outcome", 12), ("Notes", 28),
    ]
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title = ws["A1"]
    title.value     = "US Stocks — Paper Trading Journal"
    title.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    title.fill      = PatternFill("solid", start_color="0D2137")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for i, (h, w) in enumerate(headers, 1):
        _hdr_cell(ws, 2, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def _build_fee_sheet(ws):
    headers = [
        ("Date", 12), ("Instrument", 13), ("Contracts", 11),
        ("CME Fee/Leg ($)", 16), ("Broker Fee/Leg ($)", 18),
        ("NFA Fee/Leg ($)", 16), ("Total/Leg ($)", 14),
        ("Round-Trip Total ($)", 20), ("Notes", 30),
    ]
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title = ws["A1"]
    title.value     = "Fee Breakdown — All Instruments"
    title.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    title.fill      = PatternFill("solid", start_color="0D2137")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for i, (h, w) in enumerate(headers, 1):
        _hdr_cell(ws, 2, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def _build_summary_sheet(ws):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = "PAPER TRADING — PERFORMANCE SUMMARY"
    t.font      = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    t.fill      = PatternFill("solid", start_color="0D2137")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    def sub(row, text):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value     = text
        c.font      = Font(bold=True, color="1F4E79", name="Arial", size=11)
        c.fill      = SUBHDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    def lbl(row, col, text, bold=False):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=bold, name="Arial", size=10)
        c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
        c.border    = THIN
        c.fill      = WHITE_FILL

    sub(2, "NQ — Bull Call Spread")
    for r, lbl_txt in enumerate([
        "Total Trades", "Wins", "Losses", "Flat",
        "Win Rate (%)", "Gross P&L ($)", "Total Fees ($)", "Net P&L ($)",
        "Budget ($)", "Return on Budget (%)",
    ], 3):
        lbl(r, 1, lbl_txt, bold=True)
        lbl(r, 2, f"=COUNTA(NQ_Trades!A3:A1000)" if lbl_txt == "Total Trades" else "")
        lbl(r, 3, "NQ"); lbl(r, 4, "")

    sub(14, "MNQ — Bull Call Spread")
    for r, lbl_txt in enumerate([
        "Total Trades", "Wins", "Losses", "Flat",
        "Win Rate (%)", "Gross P&L ($)", "Total Fees ($)", "Net P&L ($)",
        "Budget ($)", "Return on Budget (%)",
    ], 15):
        lbl(r, 1, lbl_txt, bold=True)
        lbl(r, 2, ""); lbl(r, 3, "MNQ"); lbl(r, 4, "")

    sub(26, "Combined Totals")
    for r, lbl_txt in enumerate([
        "Total Net P&L ($)", "Total Fees Paid ($)", "Overall Win Rate (%)",
    ], 27):
        lbl(r, 1, lbl_txt, bold=True)
        lbl(r, 2, ""); lbl(r, 3, ""); lbl(r, 4, "")

    ws["B3"]  = f"=COUNTA(NQ_Trades!A3:A1000)"
    ws["B4"]  = f'=COUNTIF(NQ_Trades!V3:V1000,"PROFIT")'
    ws["B5"]  = f'=COUNTIF(NQ_Trades!V3:V1000,"STOP LOSS")'
    ws["B6"]  = f'=COUNTIF(NQ_Trades!V3:V1000,"FLAT")'
    ws["B7"]  = f'=IFERROR(B4/B3*100,0)'
    ws["B8"]  = f"=IFERROR(SUM(NQ_Trades!P3:P1000),0)"
    ws["B9"]  = f"=IFERROR(SUM(NQ_Trades!T3:T1000),0)"
    ws["B10"] = f"=IFERROR(SUM(NQ_Trades!U3:U1000),0)"
    ws["B11"] = 1000
    ws["B12"] = f"=IFERROR(B10/B11*100,0)"

    ws["B15"] = f"=COUNTA(MNQ_Trades!A3:A1000)"
    ws["B16"] = f'=COUNTIF(MNQ_Trades!V3:V1000,"PROFIT")'
    ws["B17"] = f'=COUNTIF(MNQ_Trades!V3:V1000,"STOP LOSS")'
    ws["B18"] = f'=COUNTIF(MNQ_Trades!V3:V1000,"FLAT")'
    ws["B19"] = f'=IFERROR(B16/B15*100,0)'
    ws["B20"] = f"=IFERROR(SUM(MNQ_Trades!P3:P1000),0)"
    ws["B21"] = f"=IFERROR(SUM(MNQ_Trades!T3:T1000),0)"
    ws["B22"] = f"=IFERROR(SUM(MNQ_Trades!U3:U1000),0)"
    ws["B23"] = 1000
    ws["B24"] = f"=IFERROR(B22/B23*100,0)"

    ws["B27"] = "=B10+B22"
    ws["B28"] = "=B9+B21"
    ws["B29"] = f'=IFERROR((B4+B16)/(B3+B15)*100,0)'

    for row in range(3, 30):
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            if not c.border or not c.border.left.style:
                c.border = THIN
            if not c.fill or c.fill.patternType is None:
                c.fill = WHITE_FILL


def append_trade(wb, trade: dict):
    """Append one trade row to the correct instrument sheet."""
    sheet_name = f"{trade['instrument']}_Trades"
    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    outcome = trade["outcome"]
    for i, (_, key, _) in enumerate(TRADE_COLS, 1):
        _data_cell(ws, next_row, i, trade[key], outcome)


def append_fee_row(wb, trade: dict):
    ws = wb["Fee_Breakdown"]
    next_row = ws.max_row + 1
    fees = calc_fees(INSTRUMENTS[trade["instrument"]], trade["contracts"])
    values = [
        trade["date"], trade["instrument"], trade["contracts"],
        fees["cme_fee_per_leg"], fees["broker_fee_per_leg"],
        fees["nfa_fee_per_leg"], fees["total_per_leg"],
        fees["round_trip_total"],
        f"{trade['instrument']} Bull Call Spread — {trade['outcome']}",
    ]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.font      = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border    = THIN
        c.fill      = WHITE_FILL


# ── Market day check ──────────────────────────────────────────────────────────
US_HOLIDAYS_2026 = {
    datetime.date(2026, 1,  1),   # New Year's Day
    datetime.date(2026, 1, 19),   # MLK Day
    datetime.date(2026, 2, 16),   # Presidents' Day
    datetime.date(2026, 4,  3),   # Good Friday
    datetime.date(2026, 5, 25),   # Memorial Day
    datetime.date(2026, 7,  3),   # Independence Day (observed)
    datetime.date(2026, 9,  7),   # Labor Day
    datetime.date(2026, 11, 26),  # Thanksgiving
    datetime.date(2026, 11, 27),  # Black Friday (early close — skip)
    datetime.date(2026, 12, 25),  # Christmas
}


def is_market_day(dt: datetime.date) -> bool:
    return dt.weekday() < 5 and dt not in US_HOLIDAYS_2026


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("  NQ / MNQ PAPER TRADING BOT — STARTING")
    log.info("=" * 60)

    today = datetime.date.today()
    if not is_market_day(today):
        log.info(f"{today} is not a US trading day. Exiting.")
        return

    wb = ensure_journal()
    all_trades = []

    for inst_name, inst in INSTRUMENTS.items():
        log.info(f"\n{'─'*50}")
        log.info(f"  Processing {inst_name}")
        log.info(f"{'─'*50}")

        price = fetch_price(inst["ticker"])
        if price is None:
            log.warning(f"  Could not fetch price for {inst_name}. Skipping.")
            continue

        log.info(f"  Spot price : {price:,.2f}")

        trade = simulate_trade(inst_name, inst, price)
        all_trades.append(trade)

        log.info(f"  Strategy   : {trade['strategy']}")
        log.info(f"  Strikes    : {trade['buy_strike']} / {trade['sell_strike']} CE")
        log.info(f"  Contracts  : {trade['contracts']}")
        log.info(f"  Entry Prem : ${trade['entry_premium']:,.2f}")
        log.info(f"  Exit Prem  : ${trade['exit_premium']:,.2f}")
        log.info(f"  Gross P&L  : ${trade['gross_pnl']:,.2f}")
        log.info(f"  Total Fees : ${trade['total_fees']:,.2f}")
        log.info(f"  Net P&L    : ${trade['net_pnl']:,.2f}")
        log.info(f"  Outcome    : {trade['outcome']}")

        append_trade(wb, trade)
        append_fee_row(wb, trade)

    wb.save(JOURNAL_FILE)
    log.info(f"\n✅ Journal saved → {JOURNAL_FILE}")

    # Print summary
    log.info("\n" + "=" * 60)
    log.info("  TRADE SUMMARY")
    log.info("=" * 60)
    total_net = sum(t["net_pnl"] for t in all_trades)
    total_fees = sum(t["total_fees"] for t in all_trades)
    for t in all_trades:
        log.info(f"  {t['instrument']:5s}  {t['outcome']:10s}  Net P&L: ${t['net_pnl']:>8,.2f}  Fees: ${t['total_fees']:.2f}")
    log.info(f"  {'─'*46}")
    log.info(f"  {'TOTAL':16s}           Net P&L: ${total_net:>8,.2f}  Fees: ${total_fees:.2f}")
    log.info("=" * 60)
    log.info("  Today's run complete. Next run tomorrow at 10:00 ET.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
